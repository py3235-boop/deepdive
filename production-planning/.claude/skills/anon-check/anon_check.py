# -*- coding: utf-8 -*-
"""
익명화 검사기 — 대회 산출물(코드·데이터·문서·알림 문구)에 실명·회사명·사내 식별자가 섞였는지 검사.

사용법
  python .claude/skills/anon-check/anon_check.py              # 프로젝트 전체 검사 (기본 대상 폴더)
  python .claude/skills/anon-check/anon_check.py 경로 [경로…]  # 지정 파일/폴더만 검사
  python .claude/skills/anon-check/anon_check.py --hook       # Claude Code PostToolUse 훅: stdin JSON의 file_path 1개 검사

종료 코드
  0 = 위반 없음 / 2 = 위반 발견 (훅에서는 2가 "Claude에게 stderr를 보여주고 수정하게 함")

금지어 목록은 같은 폴더의 terms.local.json 에 둔다(저장소 미포함). 패턴은 아래 PATTERNS를 고친다.
이 파일 자체는 검사 대상에서 제외된다.
"""
import io
import json
import os
import re
import sys

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

# ── 1. 금지어 (대소문자 무시) ──────────────────────────────────────────────
# 실제 고객사명·공정명·실제 품목코드·회사 도메인. 필요하면 여기에 추가.
# ── 1. 금지어 (대소문자 무시) ──────────────────────────────────────────────
# 금지어 목록은 **이 파일에 두지 않는다** — 목록 자체가 보호하려는 정보라
# 저장소에 올라가면 익명화가 통째로 무의미해지기 때문이다.
# 같은 폴더의 terms.local.json 에서 읽는다 (git 추적 제외. 형식은 terms.example.json 참고).
SELF = os.path.normpath(os.path.abspath(__file__))
TERMS_FILE = os.path.join(os.path.dirname(SELF), "terms.local.json")


def load_terms():
    """금지어 목록을 읽는다. 없으면 '검사 못 함'을 분명히 알리고 멈춘다 —
    목록 없이 통과로 처리하면 검사했다는 착각만 남기 때문이다."""
    if not os.path.exists(TERMS_FILE):
        sys.stderr.write(
            "익명화 검사를 할 수 없습니다 — 금지어 목록이 없습니다.\n"
            "  %s\n"
            "  terms.example.json 을 terms.local.json 으로 복사한 뒤 실제 값을 채우세요.\n"
            "  (이 파일은 저장소에 올리지 않습니다)\n" % TERMS_FILE)
        sys.exit(2)
    with io.open(TERMS_FILE, encoding="utf-8") as f:
        cfg = json.load(f)
    return cfg.get("terms", []), cfg.get("email_domain_regex", "")


TERMS, EMAIL_DOMAIN_RE = load_terms()

# ── 2. 정규식 패턴 (식별자·비밀값) ───────────────────────────────────────────
PATTERNS = {
    "Drive 폴더 URL":        re.compile(r"drive\.google\.com/drive/(?:u/\d+/)?folders/[A-Za-z0-9_-]{20,}"),
    "스프레드시트 URL":       re.compile(r"docs\.google\.com/spreadsheets/d/[A-Za-z0-9_-]{20,}"),
    "GAS 스크립트 URL":       re.compile(r"script\.google\.com/(?:home/)?(?:d|projects)/[A-Za-z0-9_-]{30,}"),
    "Google Chat 웹훅 URL":   re.compile(r"chat\.googleapis\.com/v1/spaces/[A-Za-z0-9_-]+/messages\?key="),
    "텔레그램 봇 토큰":       re.compile(r"\b\d{8,10}:[A-Za-z0-9_-]{35}\b"),
    "회사 이메일":            re.compile(r"[A-Za-z0-9._%+-]+@" + EMAIL_DOMAIN_RE, re.I),
}

# ── 3. 검사 대상 ────────────────────────────────────────────────────────────
DEFAULT_TARGETS = ["split", "data", "scripts", "docs", ".claude/agents", "."]   # "." 는 루트의 파일만(비재귀)
EXTS = {".gs", ".js", ".json", ".md", ".txt", ".py", ".csv", ".html"}
XLSX_EXTS = {".xlsx"}   # 셀 값 검사(openpyxl 필요 — 없으면 경고 후 생략)
EXCLUDE_DIRS = {".git", "node_modules", "__pycache__", ".claude/skills/anon-check"}

def iter_files(paths):
    for p in paths:
        p = os.path.normpath(p)
        if os.path.isfile(p):
            yield p
            continue
        if not os.path.isdir(p):
            continue
        nonrecursive = (p == "." or p == os.path.normpath(os.getcwd()))
        for root, dirs, files in os.walk(p):
            rel_root = os.path.relpath(root, ".").replace("\\", "/")
            dirs[:] = [d for d in dirs
                       if d not in EXCLUDE_DIRS
                       and (rel_root + "/" + d).lstrip("./") not in EXCLUDE_DIRS]
            for f in files:
                if f.startswith("~$"):          # Office 임시 잠금 파일
                    continue
                fp = os.path.join(root, f)
                if os.path.splitext(f)[1].lower() in EXTS | XLSX_EXTS:
                    yield fp
            if nonrecursive:
                break


def scan_xlsx(fp):
    try:
        from openpyxl import load_workbook
    except ImportError:
        print(f"⚠ openpyxl 없음 — xlsx 검사 생략: {fp}", file=sys.stderr)
        return []
    try:
        wb = load_workbook(fp, read_only=True, data_only=True)
    except Exception as e:
        print(f"⚠ xlsx 열기 실패 — 검사 생략: {fp} ({e})", file=sys.stderr)
        return []
    hits = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                s = str(cell.value)
                low = s.lower()
                for t in TERMS:
                    if t.lower() in low:
                        hits.append((cell.row, f"금지어 '{t}' [{ws.title}!{cell.coordinate}]", s.strip()[:120]))
                for name, rx in PATTERNS.items():
                    if rx.search(s):
                        hits.append((cell.row, f"{name} [{ws.title}!{cell.coordinate}]", s.strip()[:120]))
    wb.close()
    return hits


def scan_file(fp):
    if os.path.normpath(os.path.abspath(fp)) == SELF:
        return []
    if os.path.splitext(fp)[1].lower() in XLSX_EXTS:
        return scan_xlsx(fp)
    try:
        text = io.open(fp, encoding="utf-8", errors="replace").read()
    except OSError:
        return []
    hits = []
    for no, line in enumerate(text.splitlines(), 1):
        low = line.lower()
        for t in TERMS:
            if t.lower() in low:
                hits.append((no, f"금지어 '{t}'", line.strip()[:120]))
        for name, rx in PATTERNS.items():
            if rx.search(line):
                hits.append((no, name, line.strip()[:120]))
    return hits


def report(results):
    total = sum(len(h) for h in results.values())
    if total == 0:
        print("✅ 익명화 검사 통과 — 위반 0건")
        return 0
    out = [f"❌ 익명화 위반 {total}건 — 고객사A/B/C·집합XX호기 표기로 바꾸고 식별자·토큰은 스크립트 속성/CFG 플레이스홀더로 옮길 것"]
    for fp, hits in results.items():
        out.append(f"\n[{fp}]")
        for no, what, snippet in hits[:30]:
            out.append(f"  L{no:<5} {what:<18} | {snippet}")
        if len(hits) > 30:
            out.append(f"  … 외 {len(hits) - 30}건")
    msg = "\n".join(out)
    print(msg, file=sys.stderr)
    return 2


def main(argv):
    if argv and argv[0] == "--hook":
        try:
            payload = json.loads(sys.stdin.buffer.read().decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return 0
        fp = (payload.get("tool_input") or {}).get("file_path")
        if not fp or not os.path.isfile(fp):
            return 0
        if os.path.splitext(fp)[1].lower() not in EXTS | XLSX_EXTS:
            return 0
        hits = scan_file(fp)
        return report({fp: hits} if hits else {})

    targets = argv if argv else DEFAULT_TARGETS
    results = {}
    seen = set()
    for fp in iter_files(targets):
        key = os.path.normpath(os.path.abspath(fp))
        if key in seen:
            continue
        seen.add(key)
        hits = scan_file(fp)
        if hits:
            try:
                label = os.path.relpath(fp, ".")
            except ValueError:       # 다른 드라이브의 절대경로는 relpath 불가
                label = fp
            results[label] = hits
    return report(results)


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
